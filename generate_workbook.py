#!/usr/bin/env python3
"""
Excel Workbook Generator for Flash Lighting Company
Creates a comprehensive bookkeeping and accounting workbook with interconnected sheets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
from datetime import datetime
import os

# ============================================================================
# COMPANY BRANDING CONFIGURATION
# ============================================================================
# Customize these settings for your company

COMPANY_NAME = "Flash Illumination Services Ltd"
COMPANY_TAGLINE = "Professional Event Lighting Solutions"
COMPANY_EMAIL = "flashillumination.com"
COMPANY_PHONE = "+234 7032728785"
COMPANY_ADDRESS = "Lekki Gardens PhaseII, Lagos, Nigeria"
COMPANY_WEBSITE = "www.flashillumination.com"
COMPANY_IG = "@flashillumination"

# Logo Configuration
# Place your logo file (PNG, JPG, or GIF) in the same folder as this script
# and update the filename below
LOGO_FILENAME = "flash21-1.jpg"  # Change this to your logo filename
LOGO_WIDTH = 140  # Width in pixels (adjust as needed)
LOGO_HEIGHT = 80  # Height in pixels (adjust as needed)

# Color Scheme (you can customize these hex colors)
BRAND_COLOR_PRIMARY = "203864"  # Dark blue
BRAND_COLOR_SECONDARY = "366092"  # Medium blue
BRAND_COLOR_ACCENT = "4472C4"  # Light blue

# ============================================================================

def add_logo_to_sheet(ws, cell_position='A1', logo_path=None):
    """Add company logo to a worksheet if logo file exists"""
    if logo_path is None:
        logo_path = LOGO_FILENAME
    
    # Check if logo file exists
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            # Resize logo
            img.width = LOGO_WIDTH
            img.height = LOGO_HEIGHT
            # Add to worksheet
            ws.add_image(img, cell_position)
            return True
        except Exception as e:
            print(f"⚠️  Could not add logo: {e}")
            return False
    return False

def add_company_header(ws, start_row=1, include_logo=True):
    """Add standardized company header to sheets"""
    # Add logo if available
    logo_added = False
    if include_logo:
        logo_added = add_logo_to_sheet(ws, f'A{start_row}')
    
    # Adjust starting column if logo was added
    text_col = 'C' if logo_added else 'A'
    
    # Company Name
    ws[f'{text_col}{start_row}'] = COMPANY_NAME
    ws[f'{text_col}{start_row}'].font = Font(bold=True, size=16, color=BRAND_COLOR_PRIMARY)
    
    # Tagline
    ws[f'{text_col}{start_row+1}'] = COMPANY_TAGLINE
    ws[f'{text_col}{start_row+1}'].font = Font(size=10, italic=True, color="666666")
    
    # Contact info
    ws[f'{text_col}{start_row+2}'] = f"{COMPANY_PHONE} | {COMPANY_EMAIL}"
    ws[f'{text_col}{start_row+2}'].font = Font(size=9, color="666666")
    
    return start_row + 4  # Return next available row

def create_workbook():
    """Create and configure the complete workbook"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all sheets in order
    create_settings_sheet(wb)
    create_customers_sheet(wb)
    create_vendors_sheet(wb)
    create_invoice_sheet(wb)
    create_revenue_sheet(wb)
    create_expenses_sheet(wb)
    create_inventory_sheet(wb)
    create_trade_debtors_sheet(wb)
    create_cash_position_sheet(wb)
    create_bank_reconciliation_sheet(wb)
    create_profit_loss_sheet(wb)
    create_balance_sheet_sheet(wb)
    create_tax_summary_sheet(wb)
    create_invoice_template_sheet(wb)
    create_audit_log_sheet(wb)
    create_dashboard_sheet(wb)
    
    # Hide Settings and Audit Log sheets
    wb['Settings'].sheet_state = 'hidden'
    wb['Audit Log'].sheet_state = 'hidden'
    
    # Set active sheet to Dashboard
    wb.active = wb['Dashboard']
    
    return wb


def apply_header_style(ws, row, start_col, end_col):
    """Apply consistent header styling"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def create_settings_sheet(wb):
    """Create hidden Settings sheet for configuration"""
    ws = wb.create_sheet("Settings", 0)
    
    # Company Information Header
    ws['A1'] = "COMPANY SETTINGS & CONFIGURATION"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=BRAND_COLOR_PRIMARY, end_color=BRAND_COLOR_PRIMARY, fill_type="solid")
    ws.merge_cells('A1:B1')
    
    # Company Details Section
    ws['A3'] = "Company Name"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = COMPANY_NAME
    
    ws['A4'] = "Tagline"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = COMPANY_TAGLINE
    
    ws['A5'] = "Email"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = COMPANY_EMAIL
    
    ws['A6'] = "Phone"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = COMPANY_PHONE
    
    ws['A7'] = "Address"
    ws['A7'].font = Font(bold=True)
    ws['B7'] = COMPANY_ADDRESS
    
    ws['A8'] = "Website"
    ws['A8'].font = Font(bold=True)
    ws['B8'] = COMPANY_WEBSITE
    
    ws['A9'] = "Instagram"
    ws['A9'].font = Font(bold=True)
    ws['B9'] = COMPANY_IG
    
    ws['A10'] = "Logo Filename"
    ws['A10'].font = Font(bold=True)
    ws['B10'] = LOGO_FILENAME
    ws['B10'].font = Font(italic=True, color="666666")
    
    # Financial Settings
    ws['A12'] = "FINANCIAL SETTINGS"
    ws['A12'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A12'].fill = PatternFill(start_color=BRAND_COLOR_SECONDARY, end_color=BRAND_COLOR_SECONDARY, fill_type="solid")
    ws.merge_cells('A12:B12')
    
    ws['A13'] = "VAT Rate (%)"
    ws['A13'].font = Font(bold=True)
    ws['B13'] = 15
    ws['B13'].number_format = '0.00'
    
    ws['A14'] = "Financial Year Start"
    ws['A14'].font = Font(bold=True)
    ws['B14'] = "2025-01-01"
    
    # Opening Balances
    ws['A16'] = "OPENING BALANCES"
    ws['A16'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A16'].fill = PatternFill(start_color=BRAND_COLOR_SECONDARY, end_color=BRAND_COLOR_SECONDARY, fill_type="solid")
    ws.merge_cells('A16:B16')
    
    ws['A17'] = "Cash in Hand (Opening)"
    ws['A17'].font = Font(bold=True)
    ws['B17'] = 5000
    ws['B17'].number_format = '#,##0.00'
    
    ws['A18'] = "Cash at Bank (Opening)"
    ws['A18'].font = Font(bold=True)
    ws['B18'] = 50000
    ws['B18'].number_format = '#,##0.00'
    
    # Expense Categories
    ws['A20'] = "EXPENSE CATEGORIES"
    ws['A20'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A20'].fill = PatternFill(start_color=BRAND_COLOR_SECONDARY, end_color=BRAND_COLOR_SECONDARY, fill_type="solid")
    ws.merge_cells('A20:B20')
    categories = [
        "Equipment Purchase",
        "Equipment Maintenance",
        "Transport/Fuel",
        "Salaries/Wages",
        "Rent",
        "Utilities",
        "Insurance",
        "Marketing",
        "Office Supplies",
        "Professional Fees",
        "Bank Charges",
        "Other"
    ]
    for idx, cat in enumerate(categories, start=21):
        ws[f'A{idx}'] = cat
    
    # Payment Methods
    ws['D20'] = "PAYMENT METHODS"
    ws['D20'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['D20'].fill = PatternFill(start_color=BRAND_COLOR_SECONDARY, end_color=BRAND_COLOR_SECONDARY, fill_type="solid")
    ws.merge_cells('D20:E20')
    
    payment_methods = ["Cash", "Bank Transfer", "Cheque", "Mobile Money", "Card"]
    for idx, method in enumerate(payment_methods, start=21):
        ws[f'D{idx}'] = method
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['D'].width = 20


def create_customers_sheet(wb):
    """Create Customers database sheet"""
    ws = wb.create_sheet("Customers")
    
    # Headers
    headers = [
        "Customer ID", "Company Name", "Contact Person", "Email", 
        "Phone", "Address", "City", "Payment Terms (Days)", 
        "Credit Limit", "Total Invoiced", "Total Paid", "Balance", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Sample customers
    customers = [
        ("CUST-001", "Sample Client Ltd", "John Doe", "john@sampleclient.com", 
         "+234 800 123 4567", "123 Business St", "Lagos", 30, 100000),
        ("CUST-002", "Event Masters Inc", "Jane Smith", "jane@eventmasters.com", 
         "+234 800 234 5678", "456 Corporate Ave", "Abuja", 30, 150000),
        ("CUST-003", "Wedding Bliss Co", "Bob Wilson", "bob@weddingbliss.com", 
         "+234 800 345 6789", "789 Celebration Rd", "Port Harcourt", 15, 75000),
    ]
    
    for idx, (cust_id, company, contact, email, phone, address, city, terms, limit) in enumerate(customers, start=2):
        ws[f'A{idx}'] = cust_id
        ws[f'B{idx}'] = company
        ws[f'C{idx}'] = contact
        ws[f'D{idx}'] = email
        ws[f'E{idx}'] = phone
        ws[f'F{idx}'] = address
        ws[f'G{idx}'] = city
        ws[f'H{idx}'] = terms
        ws[f'I{idx}'] = limit
        ws[f'I{idx}'].number_format = '#,##0.00'
        
        # Total Invoiced
        ws[f'J{idx}'] = f'=SUMIF(Invoices!$C:$C,B{idx},Invoices!$K:$K)'
        ws[f'J{idx}'].number_format = '#,##0.00'
        
        # Total Paid
        ws[f'K{idx}'] = f'=SUMIF(Invoices!$C:$C,B{idx},Invoices!$L:$L)'
        ws[f'K{idx}'].number_format = '#,##0.00'
        
        # Balance
        ws[f'L{idx}'] = f'=J{idx}-K{idx}'
        ws[f'L{idx}'].number_format = '#,##0.00'
        
        # Status
        ws[f'M{idx}'] = f'=IF(L{idx}=0,"Paid",IF(L{idx}>I{idx}*0.8,"Credit Warning","Active"))'
    
    # Conditional formatting for status
    ws.conditional_formatting.add(f'M2:M{len(customers)+1}',
        CellIsRule(operator='equal', formula=['"Credit Warning"'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    ws.conditional_formatting.add(f'M2:M{len(customers)+1}',
        CellIsRule(operator='equal', formula=['"Paid"'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    
    # Set column widths
    widths = [12, 25, 20, 25, 18, 25, 15, 18, 15, 15, 15, 15, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A2'


def create_vendors_sheet(wb):
    """Create Vendors/Suppliers database sheet"""
    ws = wb.create_sheet("Vendors")
    
    # Headers
    headers = [
        "Vendor ID", "Company Name", "Contact Person", "Email", 
        "Phone", "Address", "City", "Payment Terms", 
        "Total Purchased", "Total Paid", "Balance Owed", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Sample vendors
    vendors = [
        ("VEND-001", "Tech Supplies Ltd", "Michael Brown", "sales@techsupplies.com", 
         "+234 800 111 2222", "15 Industrial Ave", "Lagos", "Net 30"),
        ("VEND-002", "Audio Pro Equipment", "Sarah Johnson", "info@audiopro.com", 
         "+234 800 222 3333", "28 Sound Street", "Abuja", "Net 15"),
        ("VEND-003", "Lighting World", "David Lee", "sales@lightingworld.com", 
         "+234 800 333 4444", "42 Lamp Lane", "Lagos", "Net 30"),
    ]
    
    for idx, (vend_id, company, contact, email, phone, address, city, terms) in enumerate(vendors, start=2):
        ws[f'A{idx}'] = vend_id
        ws[f'B{idx}'] = company
        ws[f'C{idx}'] = contact
        ws[f'D{idx}'] = email
        ws[f'E{idx}'] = phone
        ws[f'F{idx}'] = address
        ws[f'G{idx}'] = city
        ws[f'H{idx}'] = terms
        
        # Total Purchased
        ws[f'I{idx}'] = f'=SUMIF(Expenses!$C:$C,B{idx},Expenses!$E:$E)'
        ws[f'I{idx}'].number_format = '#,##0.00'
        
        # For now, assume all paid (can be enhanced with payables tracking)
        ws[f'J{idx}'] = f'=I{idx}'
        ws[f'J{idx}'].number_format = '#,##0.00'
        
        # Balance Owed
        ws[f'K{idx}'] = f'=I{idx}-J{idx}'
        ws[f'K{idx}'].number_format = '#,##0.00'
        
        # Status
        ws[f'L{idx}'] = f'=IF(K{idx}=0,"Current","Payable")'
    
    # Set column widths
    widths = [12, 25, 20, 25, 18, 25, 15, 15, 15, 15, 15, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A2'


def create_invoice_sheet(wb):
    """Create Invoice Generator sheet"""
    ws = wb.create_sheet("Invoices")
    
    # Headers
    headers = [
        "Invoice No.", "Date", "Client Name", "Event Date", 
        "Items/Equipment", "Quantity", "Unit Price", "Subtotal",
        "VAT Rate", "VAT Amount", "Total Amount", "Amount Received", "Balance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Sample data row 2 with formulas (template)
    ws['A2'] = "INV-2025-001"
    ws['B2'] = datetime(2025, 1, 15)
    ws['B2'].number_format = 'YYYY-MM-DD'
    ws['C2'] = "Sample Client Ltd"
    ws['D2'] = datetime(2025, 1, 20)
    ws['D2'].number_format = 'YYYY-MM-DD'
    ws['E2'] = "LED Stage Lights x10, Sound System"
    ws['F2'] = 1
    ws['G2'] = 15000
    ws['G2'].number_format = '#,##0.00'
    
    # Formulas
    ws['H2'] = "=F2*G2"  # Subtotal
    ws['H2'].number_format = '#,##0.00'
    
    ws['I2'] = "=Settings!$B$13/100"  # VAT Rate from Settings
    ws['I2'].number_format = '0.00%'
    
    ws['J2'] = "=H2*I2"  # VAT Amount
    ws['J2'].number_format = '#,##0.00'
    
    ws['K2'] = "=H2+J2"  # Total Amount
    ws['K2'].number_format = '#,##0.00'
    
    ws['L2'] = 15000  # Amount Received
    ws['L2'].number_format = '#,##0.00'
    
    ws['M2'] = "=K2-L2"  # Balance
    ws['M2'].number_format = '#,##0.00'
    
    # Data validation for Client Name (from Customers sheet)
    customer_validation = DataValidation(type="list", formula1='=Customers!$B$2:$B$100', allow_blank=False)
    customer_validation.error = 'Please select a valid customer from the list'
    customer_validation.errorTitle = 'Invalid Customer'
    ws.add_data_validation(customer_validation)
    customer_validation.add('C2:C1000')
    
    # Data validation for Payment Status (custom list)
    status_validation = DataValidation(type="list", formula1='"Paid,Partially Paid,Unpaid"', allow_blank=False)
    ws.add_data_validation(status_validation)
    status_validation.add('N2:N1000')  # Add a status column
    
    # Add Status column header
    ws['N1'] = 'Payment Status'
    ws['N1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['N1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['N1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Status formula
    ws['N2'] = '=IF(M2=0,"Paid",IF(L2>0,"Partially Paid","Unpaid"))'
    
    # Conditional formatting for Balance column
    ws.conditional_formatting.add('M2:M1000',
        CellIsRule(operator='greaterThan', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    ws.conditional_formatting.add('M2:M1000',
        CellIsRule(operator='equal', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    
    # Set column widths
    widths = [15, 12, 25, 12, 35, 10, 12, 12, 10, 12, 12, 15, 12, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'


def create_revenue_sheet(wb):
    """Create Cash Inflow / Revenue sheet"""
    ws = wb.create_sheet("Revenue")
    
    # Headers
    headers = [
        "Date", "Invoice No.", "Client Name", "Service/Equipment Provided",
        "Quantity", "Unit Price", "Total Amount", "Amount Received", "Balance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Link to Invoices sheet - Row 2 pulls from Invoices
    ws['A2'] = "=Invoices!B2"
    ws['A2'].number_format = 'YYYY-MM-DD'
    ws['B2'] = "=Invoices!A2"
    ws['C2'] = "=Invoices!C2"
    ws['D2'] = "=Invoices!E2"
    ws['E2'] = "=Invoices!F2"
    ws['F2'] = "=Invoices!G2"
    ws['F2'].number_format = '#,##0.00'
    ws['G2'] = "=Invoices!K2"
    ws['G2'].number_format = '#,##0.00'
    ws['H2'] = "=Invoices!L2"
    ws['H2'].number_format = '#,##0.00'
    ws['I2'] = "=Invoices!M2"
    ws['I2'].number_format = '#,##0.00'
    
    # Summary section
    ws['K1'] = "REVENUE SUMMARY"
    ws['K1'].font = Font(bold=True, size=12)
    
    ws['K3'] = "Today's Revenue"
    ws['L3'] = f'=SUMIFS(H:H,A:A,TODAY())'
    ws['L3'].number_format = '#,##0.00'
    
    ws['K4'] = "This Month's Revenue"
    ws['L4'] = f'=SUMIFS(H:H,A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
    ws['L4'].number_format = '#,##0.00'
    
    ws['K5'] = "This Year's Revenue"
    ws['L5'] = f'=SUMIFS(H:H,A:A,">="&DATE(YEAR(TODAY()),1,1),A:A,"<"&DATE(YEAR(TODAY())+1,1,1))'
    ws['L5'].number_format = '#,##0.00'
    
    ws['K7'] = "Total Outstanding"
    ws['L7'] = '=SUM(I:I)'
    ws['L7'].number_format = '#,##0.00'
    
    # Set column widths
    widths = [12, 15, 25, 35, 10, 12, 12, 15, 12, 2, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A2'


def create_expenses_sheet(wb):
    """Create Expenses sheet"""
    ws = wb.create_sheet("Expenses")
    
    # Headers
    headers = [
        "Date", "Expense Category", "Vendor", "Description", 
        "Amount", "Payment Method", "Reference"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Sample data
    ws['A2'] = datetime(2025, 1, 10)
    ws['A2'].number_format = 'YYYY-MM-DD'
    ws['B2'] = "Equipment Purchase"
    ws['C2'] = "Tech Supplies Ltd"
    ws['D2'] = "10x LED Par Lights"
    ws['E2'] = 25000
    ws['E2'].number_format = '#,##0.00'
    ws['F2'] = "Bank Transfer"
    ws['G2'] = "TXN-001"
    
    # Summary section
    ws['I1'] = "EXPENSE SUMMARY"
    ws['I1'].font = Font(bold=True, size=12)
    
    ws['I3'] = "This Month's Expenses"
    ws['J3'] = f'=SUMIFS(E:E,A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),A:A,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))'
    ws['J3'].number_format = '#,##0.00'
    
    ws['I4'] = "This Year's Expenses"
    ws['J4'] = f'=SUMIFS(E:E,A:A,">="&DATE(YEAR(TODAY()),1,1),A:A,"<"&DATE(YEAR(TODAY())+1,1,1))'
    ws['J4'].number_format = '#,##0.00'
    
    ws['I6'] = "By Category (YTD)"
    ws['I6'].font = Font(bold=True)
    
    # Category summaries
    categories = [
        "Equipment Purchase", "Equipment Maintenance", "Transport/Fuel",
        "Salaries/Wages", "Rent", "Utilities", "Insurance", "Marketing",
        "Office Supplies", "Professional Fees", "Bank Charges", "Other"
    ]
    
    for idx, cat in enumerate(categories, start=7):
        ws[f'I{idx}'] = cat
        ws[f'J{idx}'] = f'=SUMIFS(E:E,B:B,I{idx},A:A,">="&DATE(YEAR(TODAY()),1,1))'
        ws[f'J{idx}'].number_format = '#,##0.00'
    
    # Data validation for Expense Category (from Settings sheet)
    category_validation = DataValidation(type="list", formula1='=Settings!$A$21:$A$32', allow_blank=False)
    category_validation.error = 'Please select a valid expense category'
    category_validation.errorTitle = 'Invalid Category'
    ws.add_data_validation(category_validation)
    category_validation.add('B2:B1000')
    
    # Data validation for Vendor (from Vendors sheet)
    vendor_validation = DataValidation(type="list", formula1='=Vendors!$B$2:$B$100', allow_blank=False)
    vendor_validation.error = 'Please select a valid vendor'
    vendor_validation.errorTitle = 'Invalid Vendor'
    ws.add_data_validation(vendor_validation)
    vendor_validation.add('C2:C1000')
    
    # Data validation for Payment Method (from Settings sheet)
    payment_validation = DataValidation(type="list", formula1='=Settings!$D$21:$D$25', allow_blank=False)
    payment_validation.error = 'Please select a valid payment method'
    payment_validation.errorTitle = 'Invalid Payment Method'
    ws.add_data_validation(payment_validation)
    payment_validation.add('F2:F1000')
    
    # Set column widths
    widths = [12, 20, 25, 35, 12, 18, 15, 2, 25, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A2'


def create_inventory_sheet(wb):
    """Create Stock / Inventory sheet"""
    ws = wb.create_sheet("Inventory")
    
    # Headers
    headers = [
        "Stock ID", "Item Description", "Unit Price", "Quantity in Store",
        "Quantity Rented Out", "Stock in Transit", "Total Quantity", "Total Stock Value"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Sample data with formulas
    inventory_items = [
        ("STK-001", "LED Par Light", 2500, 20, 5, 0),
        ("STK-002", "Moving Head Light", 8000, 10, 3, 2),
        ("STK-003", "Sound System (Complete)", 45000, 3, 1, 0),
        ("STK-004", "Smoke Machine", 3500, 8, 2, 0),
        ("STK-005", "DMX Controller", 5000, 5, 1, 0),
        ("STK-006", "Lighting Truss (per meter)", 800, 50, 20, 0),
        ("STK-007", "Power Distribution Box", 1200, 15, 5, 0),
        ("STK-008", "Microphone (Wireless)", 1500, 12, 4, 0),
        ("STK-009", "Speaker (500W)", 6000, 8, 3, 1),
        ("STK-010", "Cable Set (Complete)", 300, 30, 10, 5),
    ]
    
    for idx, (stock_id, desc, price, in_store, rented, transit) in enumerate(inventory_items, start=2):
        ws[f'A{idx}'] = stock_id
        ws[f'B{idx}'] = desc
        ws[f'C{idx}'] = price
        ws[f'C{idx}'].number_format = '#,##0.00'
        ws[f'D{idx}'] = in_store
        ws[f'E{idx}'] = rented
        ws[f'F{idx}'] = transit
        
        # Total Quantity formula
        ws[f'G{idx}'] = f'=D{idx}+E{idx}+F{idx}'
        
        # Total Stock Value formula
        ws[f'H{idx}'] = f'=C{idx}*G{idx}'
        ws[f'H{idx}'].number_format = '#,##0.00'
    
    # Summary section
    last_row = len(inventory_items) + 2
    ws[f'A{last_row}'] = "TOTAL STOCK VALUE"
    ws[f'A{last_row}'].font = Font(bold=True, size=11)
    ws[f'H{last_row}'] = f'=SUM(H2:H{last_row-1})'
    ws[f'H{last_row}'].number_format = '#,##0.00'
    ws[f'H{last_row}'].font = Font(bold=True)
    ws[f'H{last_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Additional summary
    ws[f'J3'] = "Available for Rent"
    ws[f'K3'] = f'=SUM(D:D)'
    
    ws[f'J4'] = "Currently Rented"
    ws[f'K4'] = f'=SUM(E:E)'
    
    ws[f'J5'] = "In Transit"
    ws[f'K5'] = f'=SUM(F:F)'
    
    # Set column widths
    widths = [12, 30, 12, 18, 18, 15, 15, 18, 2, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A2'


def create_trade_debtors_sheet(wb):
    """Create Trade Debtors (Accounts Receivable) sheet"""
    ws = wb.create_sheet("Trade Debtors")
    
    # Headers
    headers = [
        "Invoice No.", "Date", "Client Name", "Amount Owed", 
        "Due Date", "Days Outstanding", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
    
    apply_header_style(ws, 1, 1, len(headers))
    
    # Instructions
    ws['A2'] = "This sheet auto-populates from Invoices where Balance > 0"
    ws['A2'].font = Font(italic=True, color="666666")
    ws.merge_cells('A2:G2')
    
    # Formula to pull unpaid invoices
    # Row 4 onwards will contain the actual data
    ws['A4'] = '=IFERROR(INDEX(Invoices!$A:$A,SMALL(IF(Invoices!$M$2:$M$1000>0,ROW(Invoices!$M$2:$M$1000)),ROW()-3)),"")'
    ws['B4'] = '=IFERROR(INDEX(Invoices!$B:$B,MATCH(A4,Invoices!$A:$A,0)),"")'
    ws['B4'].number_format = 'YYYY-MM-DD'
    ws['C4'] = '=IFERROR(INDEX(Invoices!$C:$C,MATCH(A4,Invoices!$A:$A,0)),"")'
    ws['D4'] = '=IFERROR(INDEX(Invoices!$M:$M,MATCH(A4,Invoices!$A:$A,0)),"")'
    ws['D4'].number_format = '#,##0.00'
    ws['E4'] = '=IFERROR(B4+30,"")'  # Due date = Invoice date + 30 days
    ws['E4'].number_format = 'YYYY-MM-DD'
    ws['F4'] = '=IFERROR(IF(A4<>"",TODAY()-B4,""),"")'
    ws['G4'] = '=IFERROR(IF(F4="","",IF(F4>60,"Overdue",IF(F4>30,"Due Soon","Current"))),"")'
    
    # Summary
    ws['I3'] = "Total Outstanding"
    ws['J3'] = '=SUM(D:D)'
    ws['J3'].number_format = '#,##0.00'
    ws['J3'].font = Font(bold=True)
    
    ws['I5'] = "Current (0-30 days)"
    ws['J5'] = '=SUMIF(G:G,"Current",D:D)'
    ws['J5'].number_format = '#,##0.00'
    
    ws['I6'] = "Due Soon (31-60 days)"
    ws['J6'] = '=SUMIF(G:G,"Due Soon",D:D)'
    ws['J6'].number_format = '#,##0.00'
    
    ws['I7'] = "Overdue (>60 days)"
    ws['J7'] = '=SUMIF(G:G,"Overdue",D:D)'
    ws['J7'].number_format = '#,##0.00'
    
    # Set column widths
    widths = [15, 12, 25, 15, 12, 18, 15, 2, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A4'


def create_cash_position_sheet(wb):
    """Create Cash Position sheet"""
    ws = wb.create_sheet("Cash Position")
    
    # Title
    ws['A1'] = "CASH POSITION STATEMENT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"As at: {datetime.now().strftime('%Y-%m-%d')}"
    ws.merge_cells('A2:D2')
    
    # Cash in Hand section
    ws['A4'] = "CASH IN HAND"
    ws['A4'].font = Font(bold=True, size=12)
    apply_header_style(ws, 4, 1, 2)
    
    ws['A5'] = "Opening Balance"
    ws['B5'] = "=Settings!B17"
    ws['B5'].number_format = '#,##0.00'
    
    ws['A6'] = "Cash Received (YTD)"
    ws['B6'] = '=SUMIFS(Revenue!H:H,Revenue!A:A,">="&Settings!$B$14,Expenses!F:F,"Cash")'
    ws['B6'].number_format = '#,##0.00'
    
    ws['A7'] = "Cash Paid (YTD)"
    ws['B7'] = '=SUMIFS(Expenses!E:E,Expenses!A:A,">="&Settings!$B$14,Expenses!F:F,"Cash")'
    ws['B7'].number_format = '#,##0.00'
    
    ws['A8'] = "Closing Cash in Hand"
    ws['B8'] = "=B5+B6-B7"
    ws['B8'].number_format = '#,##0.00'
    ws['B8'].font = Font(bold=True)
    ws['B8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Cash at Bank section
    ws['A10'] = "CASH AT BANK"
    ws['A10'].font = Font(bold=True, size=12)
    apply_header_style(ws, 10, 1, 2)
    
    ws['A11'] = "Opening Balance"
    ws['B11'] = "=Settings!B18"
    ws['B11'].number_format = '#,##0.00'
    
    ws['A12'] = "Bank Receipts (YTD)"
    ws['B12'] = '=SUMIFS(Revenue!H:H,Revenue!A:A,">="&Settings!$B$14)-B6'
    ws['B12'].number_format = '#,##0.00'
    
    ws['A13'] = "Bank Payments (YTD)"
    ws['B13'] = '=SUMIFS(Expenses!E:E,Expenses!A:A,">="&Settings!$B$14)-B7'
    ws['B13'].number_format = '#,##0.00'
    
    ws['A14'] = "Closing Cash at Bank"
    ws['B14'] = "=B11+B12-B13"
    ws['B14'].number_format = '#,##0.00'
    ws['B14'].font = Font(bold=True)
    ws['B14'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Total Cash Position
    ws['A16'] = "TOTAL CASH POSITION"
    ws['A16'].font = Font(bold=True, size=13)
    ws['B16'] = "=B8+B14"
    ws['B16'].number_format = '#,##0.00'
    ws['B16'].font = Font(bold=True, size=13)
    ws['B16'].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    
    # Monthly Breakdown
    ws['D4'] = "MONTHLY CASH FLOW"
    ws['D4'].font = Font(bold=True, size=12)
    ws.merge_cells('D4:F4')
    
    ws['D5'] = "Month"
    ws['E5'] = "Inflows"
    ws['F5'] = "Outflows"
    ws['G5'] = "Net"
    apply_header_style(ws, 5, 4, 7)
    
    # Sample months (would need dynamic generation for full year)
    months = ["January", "February", "March", "April", "May", "June", 
              "July", "August", "September", "October", "November", "December"]
    
    for idx, month in enumerate(months, start=6):
        month_num = idx - 5
        ws[f'D{idx}'] = month
        ws[f'E{idx}'] = f'=SUMIFS(Revenue!H:H,Revenue!A:A,">="&DATE(YEAR(TODAY()),{month_num},1),Revenue!A:A,"<"&DATE(YEAR(TODAY()),{month_num}+1,1))'
        ws[f'E{idx}'].number_format = '#,##0.00'
        ws[f'F{idx}'] = f'=SUMIFS(Expenses!E:E,Expenses!A:A,">="&DATE(YEAR(TODAY()),{month_num},1),Expenses!A:A,"<"&DATE(YEAR(TODAY()),{month_num}+1,1))'
        ws[f'F{idx}'].number_format = '#,##0.00'
        ws[f'G{idx}'] = f'=E{idx}-F{idx}'
        ws[f'G{idx}'].number_format = '#,##0.00'
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15


def create_bank_reconciliation_sheet(wb):
    """Create Bank Reconciliation sheet like QuickBooks"""
    ws = wb.create_sheet("Bank Reconciliation")
    
    # Title section
    ws['A1'] = "BANK RECONCILIATION STATEMENT"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A2'] = f"As at: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Bank Statement Section
    ws['A4'] = "BANK STATEMENT BALANCE"
    ws['A4'].font = Font(bold=True, size=12)
    apply_header_style(ws, 4, 1, 2)
    
    ws['A5'] = "Bank Statement Balance"
    ws['B5'] = 50000  # Default value, user should update
    ws['B5'].number_format = '#,##0.00'
    ws['B5'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Outstanding Transactions
    ws['A7'] = "OUTSTANDING DEPOSITS (Not yet cleared)"
    ws['A7'].font = Font(bold=True, size=11)
    apply_header_style(ws, 7, 1, 3)
    
    ws['A8'] = "Date"
    ws['B8'] = "Description"
    ws['C8'] = "Amount"
    apply_header_style(ws, 8, 1, 3)
    
    # Sample outstanding deposit
    ws['A9'] = datetime.now()
    ws['A9'].number_format = 'YYYY-MM-DD'
    ws['B9'] = "Client payment in transit"
    ws['C9'] = 0
    ws['C9'].number_format = '#,##0.00'
    
    ws['A11'] = "Total Outstanding Deposits"
    ws['A11'].font = Font(bold=True)
    ws['C11'] = '=SUM(C9:C10)'
    ws['C11'].number_format = '#,##0.00'
    ws['C11'].font = Font(bold=True)
    
    # Outstanding Checks
    ws['A13'] = "OUTSTANDING CHECKS/PAYMENTS (Not yet cleared)"
    ws['A13'].font = Font(bold=True, size=11)
    apply_header_style(ws, 13, 1, 3)
    
    ws['A14'] = "Date"
    ws['B14'] = "Check No./Description"
    ws['C14'] = "Amount"
    apply_header_style(ws, 14, 1, 3)
    
    # Sample outstanding check
    ws['A15'] = datetime.now()
    ws['A15'].number_format = 'YYYY-MM-DD'
    ws['B15'] = "CHQ-001 to vendor"
    ws['C15'] = 0
    ws['C15'].number_format = '#,##0.00'
    
    ws['A17'] = "Total Outstanding Checks"
    ws['A17'].font = Font(bold=True)
    ws['C17'] = '=SUM(C15:C16)'
    ws['C17'].number_format = '#,##0.00'
    ws['C17'].font = Font(bold=True)
    
    # Reconciliation Summary
    ws['A19'] = "RECONCILIATION SUMMARY"
    ws['A19'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A19'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A19:B19')
    
    ws['A20'] = "Bank Statement Balance"
    ws['B20'] = '=B5'
    ws['B20'].number_format = '#,##0.00'
    
    ws['A21'] = "Add: Outstanding Deposits"
    ws['B21'] = '=C11'
    ws['B21'].number_format = '#,##0.00'
    
    ws['A22'] = "Less: Outstanding Checks"
    ws['B22'] = '=-C17'
    ws['B22'].number_format = '#,##0.00'
    
    ws['A23'] = "Adjusted Bank Balance"
    ws['B23'] = '=B20+B21+B22'
    ws['B23'].number_format = '#,##0.00'
    ws['B23'].font = Font(bold=True, size=12)
    ws['B23'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws['A25'] = "Book Balance (Per Cash Position)"
    ws['B25'] = "='Cash Position'!B14"
    ws['B25'].number_format = '#,##0.00'
    ws['B25'].font = Font(bold=True)
    
    ws['A27'] = "DIFFERENCE (Should be zero)"
    ws['A27'].font = Font(bold=True)
    ws['B27'] = '=B23-B25'
    ws['B27'].number_format = '#,##0.00'
    ws['B27'].font = Font(bold=True, size=12)
    
    # Conditional formatting for difference
    ws.conditional_formatting.add('B27:B27',
        CellIsRule(operator='equal', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    
    ws.conditional_formatting.add('B27:B27',
        CellIsRule(operator='notEqual', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    

def create_profit_loss_sheet(wb):
    """Create professional Profit & Loss Statement like QuickBooks"""
    ws = wb.create_sheet("Profit & Loss")
    
    # Title
    ws['A1'] = "PROFIT & LOSS STATEMENT"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"For the period: January 1, 2025 - {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:D2')
    
    # Column Headers
    ws['C4'] = "This Month"
    ws['D4'] = "Year to Date"
    apply_header_style(ws, 4, 3, 4)
    
    # REVENUE SECTION
    ws['A5'] = "REVENUE"
    ws['A5'].font = Font(bold=True, size=13, color="FFFFFF")
    ws['A5'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A5:B5')
    
    ws['A6'] = "Service Revenue"
    ws['C6'] = '=Revenue!L4'
    ws['C6'].number_format = '#,##0.00'
    ws['D6'] = '=Revenue!L5'
    ws['D6'].number_format = '#,##0.00'
    
    ws['A7'] = "Total Revenue"
    ws['A7'].font = Font(bold=True)
    ws['C7'] = '=C6'
    ws['C7'].number_format = '#,##0.00'
    ws['C7'].font = Font(bold=True)
    ws['D7'] = '=D6'
    ws['D7'].number_format = '#,##0.00'
    ws['D7'].font = Font(bold=True)
    
    # COST OF SERVICES
    ws['A9'] = "COST OF SERVICES"
    ws['A9'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A9:B9')
    
    ws['A10'] = "Equipment Costs"
    ws['C10'] = f'=SUMIFS(Expenses!E:E,Expenses!B:B,"Equipment Purchase",Expenses!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))'
    ws['C10'].number_format = '#,##0.00'
    ws['D10'] = f'=SUMIFS(Expenses!E:E,Expenses!B:B,"Equipment Purchase",Expenses!A:A,">="&Settings!$B$14)'
    ws['D10'].number_format = '#,##0.00'
    
    ws['A11'] = "Total Cost of Services"
    ws['A11'].font = Font(bold=True)
    ws['C11'] = '=C10'
    ws['C11'].number_format = '#,##0.00'
    ws['C11'].font = Font(bold=True)
    ws['D11'] = '=D10'
    ws['D11'].number_format = '#,##0.00'
    ws['D11'].font = Font(bold=True)
    
    # GROSS PROFIT
    ws['A13'] = "GROSS PROFIT"
    ws['A13'].font = Font(bold=True, size=13)
    ws['C13'] = '=C7-C11'
    ws['C13'].number_format = '#,##0.00'
    ws['C13'].font = Font(bold=True, size=12)
    ws['C13'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws['D13'] = '=D7-D11'
    ws['D13'].number_format = '#,##0.00'
    ws['D13'].font = Font(bold=True, size=12)
    ws['D13'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # OPERATING EXPENSES
    ws['A15'] = "OPERATING EXPENSES"
    ws['A15'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A15'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A15:B15')
    
    expense_categories = [
        "Equipment Maintenance", "Transport/Fuel", "Salaries/Wages", "Rent",
        "Utilities", "Insurance", "Marketing", "Office Supplies",
        "Professional Fees", "Bank Charges", "Other"
    ]
    
    start_row = 16
    for idx, cat in enumerate(expense_categories, start=start_row):
        ws[f'A{idx}'] = cat
        ws[f'C{idx}'] = f'=SUMIFS(Expenses!E:E,Expenses!B:B,A{idx},Expenses!A:A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))'
        ws[f'C{idx}'].number_format = '#,##0.00'
        ws[f'D{idx}'] = f'=SUMIFS(Expenses!E:E,Expenses!B:B,A{idx},Expenses!A:A,">="&Settings!$B$14)'
        ws[f'D{idx}'].number_format = '#,##0.00'
    
    total_row = start_row + len(expense_categories)
    ws[f'A{total_row}'] = "Total Operating Expenses"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'C{total_row}'] = f'=SUM(C{start_row}:C{total_row-1})'
    ws[f'C{total_row}'].number_format = '#,##0.00'
    ws[f'C{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = f'=SUM(D{start_row}:D{total_row-1})'
    ws[f'D{total_row}'].number_format = '#,##0.00'
    ws[f'D{total_row}'].font = Font(bold=True)
    
    # NET INCOME
    net_row = total_row + 2
    ws[f'A{net_row}'] = "NET INCOME (LOSS)"
    ws[f'A{net_row}'].font = Font(bold=True, size=14)
    ws[f'C{net_row}'] = f'=C13-C{total_row}'
    ws[f'C{net_row}'].number_format = '#,##0.00'
    ws[f'C{net_row}'].font = Font(bold=True, size=13)
    ws[f'C{net_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    ws[f'D{net_row}'] = f'=D13-D{total_row}'
    ws[f'D{net_row}'].number_format = '#,##0.00'
    ws[f'D{net_row}'].font = Font(bold=True, size=13)
    ws[f'D{net_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Profit Margin
    margin_row = net_row + 1
    ws[f'A{margin_row}'] = "Net Profit Margin"
    ws[f'A{margin_row}'].font = Font(bold=True)
    ws[f'C{margin_row}'] = f'=IF(C7=0,0,C{net_row}/C7)'
    ws[f'C{margin_row}'].number_format = '0.00%'
    ws[f'D{margin_row}'] = f'=IF(D7=0,0,D{net_row}/D7)'
    ws[f'D{margin_row}'].number_format = '0.00%'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18


def create_balance_sheet_sheet(wb):
    """Create professional Balance Sheet like QuickBooks"""
    ws = wb.create_sheet("Balance Sheet")
    
    # Title
    ws['A1'] = "BALANCE SHEET"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:C1')
    
    ws['A2'] = f"As at: {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:C2')
    
    # ASSETS
    ws['A4'] = "ASSETS"
    ws['A4'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A4:B4')
    
    ws['A5'] = "Current Assets"
    ws['A5'].font = Font(bold=True, size=12)
    
    ws['A6'] = "  Cash in Hand"
    ws['C6'] = "='Cash Position'!B8"
    ws['C6'].number_format = '#,##0.00'
    
    ws['A7'] = "  Cash at Bank"
    ws['C7'] = "='Cash Position'!B14"
    ws['C7'].number_format = '#,##0.00'
    
    ws['A8'] = "  Accounts Receivable"
    ws['C8'] = "='Trade Debtors'!J3"
    ws['C8'].number_format = '#,##0.00'
    
    ws['A9'] = "Total Current Assets"
    ws['A9'].font = Font(bold=True)
    ws['C9'] = '=SUM(C6:C8)'
    ws['C9'].number_format = '#,##0.00'
    ws['C9'].font = Font(bold=True)
    ws['C9'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['A11'] = "Fixed Assets"
    ws['A11'].font = Font(bold=True, size=12)
    
    ws['A12'] = "  Equipment & Inventory"
    ws['C12'] = '=Inventory!H12'
    ws['C12'].number_format = '#,##0.00'
    
    ws['A13'] = "Total Fixed Assets"
    ws['A13'].font = Font(bold=True)
    ws['C13'] = '=C12'
    ws['C13'].number_format = '#,##0.00'
    ws['C13'].font = Font(bold=True)
    ws['C13'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['A15'] = "TOTAL ASSETS"
    ws['A15'].font = Font(bold=True, size=13)
    ws['C15'] = '=C9+C13'
    ws['C15'].number_format = '#,##0.00'
    ws['C15'].font = Font(bold=True, size=13)
    ws['C15'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # LIABILITIES & EQUITY
    ws['A17'] = "LIABILITIES & EQUITY"
    ws['A17'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A17'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A17:B17')
    
    ws['A18'] = "Current Liabilities"
    ws['A18'].font = Font(bold=True, size=12)
    
    ws['A19'] = "  Accounts Payable"
    ws['C19'] = 0  # Can be enhanced with payables tracking
    ws['C19'].number_format = '#,##0.00'
    
    ws['A20'] = "  VAT Payable"
    ws['C20'] = "='Tax Summary'!B12"
    ws['C20'].number_format = '#,##0.00'
    
    ws['A21'] = "Total Current Liabilities"
    ws['A21'].font = Font(bold=True)
    ws['C21'] = '=SUM(C19:C20)'
    ws['C21'].number_format = '#,##0.00'
    ws['C21'].font = Font(bold=True)
    ws['C21'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['A23'] = "Owner's Equity"
    ws['A23'].font = Font(bold=True, size=12)
    
    ws['A24'] = "  Opening Capital"
    ws['C24'] = '=Settings!B17+Settings!B18'  # Opening balances
    ws['C24'].number_format = '#,##0.00'
    
    ws['A25'] = "  Retained Earnings (YTD)"
    ws['C25'] = "='Profit & Loss'!D29"  # Net income
    ws['C25'].number_format = '#,##0.00'
    
    ws['A26'] = "Total Owner's Equity"
    ws['A26'].font = Font(bold=True)
    ws['C26'] = '=C24+C25'
    ws['C26'].number_format = '#,##0.00'
    ws['C26'].font = Font(bold=True)
    ws['C26'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['A28'] = "TOTAL LIABILITIES & EQUITY"
    ws['A28'].font = Font(bold=True, size=13)
    ws['C28'] = '=C21+C26'
    ws['C28'].number_format = '#,##0.00'
    ws['C28'].font = Font(bold=True, size=13)
    ws['C28'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Balance Check
    ws['A30'] = "CHECK (Should be zero):"
    ws['A30'].font = Font(bold=True, italic=True)
    ws['C30'] = '=C15-C28'
    ws['C30'].number_format = '#,##0.00'
    ws['C30'].font = Font(bold=True)
    
    # Conditional formatting for balance check
    ws.conditional_formatting.add('C30:C30',
        CellIsRule(operator='equal', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    
    ws.conditional_formatting.add('C30:C30',
        CellIsRule(operator='notEqual', formula=['0'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20


def create_tax_summary_sheet(wb):
    """Create Tax Summary sheet for end-of-year reporting"""
    ws = wb.create_sheet("Tax Summary")
    
    # Title
    ws['A1'] = "END-OF-YEAR TAX SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Financial Year: 2025"
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A2:D2')
    
    # Revenue Section
    ws['A4'] = "REVENUE"
    ws['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A4'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A4:B4')
    
    ws['A5'] = "Total Revenue (Invoiced)"
    ws['B5'] = '=SUMIFS(Revenue!G:G,Revenue!A:A,">="&Settings!$B$14,Revenue!A:A,"<"&DATE(YEAR(Settings!$B$14)+1,1,1))'
    ws['B5'].number_format = '#,##0.00'
    
    ws['A6'] = "Total Revenue (Received)"
    ws['B6'] = '=SUMIFS(Revenue!H:H,Revenue!A:A,">="&Settings!$B$14,Revenue!A:A,"<"&DATE(YEAR(Settings!$B$14)+1,1,1))'
    ws['B6'].number_format = '#,##0.00'
    
    ws['A7'] = "Outstanding Receivables"
    ws['B7'] = "=B5-B6"
    ws['B7'].number_format = '#,##0.00'
    
    # VAT Section
    ws['A9'] = "VALUE ADDED TAX (VAT)"
    ws['A9'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A9'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A9:B9')
    
    ws['A10'] = "VAT Collected on Sales"
    ws['B10'] = '=SUMIFS(Invoices!J:J,Invoices!B:B,">="&Settings!$B$14,Invoices!B:B,"<"&DATE(YEAR(Settings!$B$14)+1,1,1))'
    ws['B10'].number_format = '#,##0.00'
    
    ws['A11'] = "VAT Paid on Purchases"
    ws['B11'] = '=SUMIFS(Expenses!E:E,Expenses!A:A,">="&Settings!$B$14,Expenses!A:A,"<"&DATE(YEAR(Settings!$B$14)+1,1,1))*(Settings!$B$13/100)/(1+Settings!$B$13/100)'
    ws['B11'].number_format = '#,##0.00'
    
    ws['A12'] = "Net VAT Payable"
    ws['B12'] = "=B10-B11"
    ws['B12'].number_format = '#,##0.00'
    ws['B12'].font = Font(bold=True)
    ws['B12'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Expenses Section
    ws['A14'] = "EXPENSES (Deductible)"
    ws['A14'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A14'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A14:B14')
    
    categories = [
        "Equipment Purchase", "Equipment Maintenance", "Transport/Fuel",
        "Salaries/Wages", "Rent", "Utilities", "Insurance", "Marketing",
        "Office Supplies", "Professional Fees", "Bank Charges", "Other"
    ]
    
    start_row = 15
    for idx, cat in enumerate(categories, start=start_row):
        ws[f'A{idx}'] = cat
        ws[f'B{idx}'] = f'=SUMIFS(Expenses!E:E,Expenses!B:B,A{idx},Expenses!A:A,">="&Settings!$B$14,Expenses!A:A,"<"&DATE(YEAR(Settings!$B$14)+1,1,1))'
        ws[f'B{idx}'].number_format = '#,##0.00'
    
    total_row = start_row + len(categories)
    ws[f'A{total_row}'] = "Total Expenses"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'] = f'=SUM(B{start_row}:B{total_row-1})'
    ws[f'B{total_row}'].number_format = '#,##0.00'
    ws[f'B{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Net Profit Section
    profit_row = total_row + 2
    ws[f'A{profit_row}'] = "NET PROFIT BEFORE TAX"
    ws[f'A{profit_row}'].font = Font(bold=True, size=13)
    ws[f'B{profit_row}'] = f'=B6-B{total_row}'
    ws[f'B{profit_row}'].number_format = '#,##0.00'
    ws[f'B{profit_row}'].font = Font(bold=True, size=13)
    ws[f'B{profit_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Profit Margin
    margin_row = profit_row + 1
    ws[f'A{margin_row}'] = "Profit Margin (%)"
    ws[f'B{margin_row}'] = f'=IF(B6=0,0,B{profit_row}/B6)'
    ws[f'B{margin_row}'].number_format = '0.00%'
    ws[f'B{margin_row}'].font = Font(bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_invoice_template_sheet(wb):
    """Create printable Invoice Template sheet like QuickBooks"""
    ws = wb.create_sheet("Invoice Template")
    
    # Add company logo
    logo_added = add_logo_to_sheet(ws, 'A1')
    
    # Company Header - adjust if logo was added
    text_col = 'C' if logo_added else 'A'
    
    # Company Name
    ws[f'{text_col}1'] = COMPANY_NAME
    ws[f'{text_col}1'].font = Font(bold=True, size=20, color=BRAND_COLOR_PRIMARY)
    if not logo_added:
        ws.merge_cells('A1:C1')
    
    # Tagline
    ws[f'{text_col}2'] = COMPANY_TAGLINE
    ws[f'{text_col}2'].font = Font(size=11, italic=True, color="666666")
    if not logo_added:
        ws.merge_cells('A2:C2')
    
    # Contact Information
    ws[f'{text_col}3'] = f"{COMPANY_PHONE} | {COMPANY_EMAIL}"
    ws[f'{text_col}3'].font = Font(size=9, color="666666")
    if not logo_added:
        ws.merge_cells('A3:C3')
    
    ws[f'{text_col}4'] = f"{COMPANY_ADDRESS} | {COMPANY_WEBSITE}"
    ws[f'{text_col}4'].font = Font(size=9, color="666666")
    if not logo_added:
        ws.merge_cells('A4:C4')
    
    # Invoice Title
    ws['E1'] = "INVOICE"
    ws['E1'].font = Font(bold=True, size=24, color="FFFFFF")
    ws['E1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('E1:F2')
    
    # Invoice Details Section
    ws['A5'] = "BILL TO:"
    ws['A5'].font = Font(bold=True, size=11)
    
    ws['A6'] = "Client Name:"
    ws['B6'] = "[Select from Invoices sheet]"
    ws['B6'].font = Font(italic=True, color="666666")
    
    ws['A7'] = "Address:"
    ws['B7'] = "[Auto-filled from Customers]"
    ws['B7'].font = Font(italic=True, color="666666")
    
    ws['A8'] = "Phone:"
    ws['B8'] = "[Auto-filled from Customers]"
    ws['B8'].font = Font(italic=True, color="666666")
    
    # Invoice Meta
    ws['E5'] = "Invoice No:"
    ws['E5'].font = Font(bold=True)
    ws['F5'] = "INV-2025-001"
    ws['F5'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws['E6'] = "Invoice Date:"
    ws['E6'].font = Font(bold=True)
    ws['F6'] = datetime.now()
    ws['F6'].number_format = 'YYYY-MM-DD'
    
    ws['E7'] = "Due Date:"
    ws['E7'].font = Font(bold=True)
    ws['F7'] = datetime.now()
    ws['F7'].number_format = 'YYYY-MM-DD'
    
    # Items Table
    ws['A10'] = "Description"
    ws['B10'] = "Quantity"
    ws['C10'] = "Unit Price"
    ws['D10'] = "Amount"
    apply_header_style(ws, 10, 1, 4)
    
    # Sample items
    ws['A11'] = "LED Stage Lights"
    ws['B11'] = 10
    ws['C11'] = 1200
    ws['C11'].number_format = '#,##0.00'
    ws['D11'] = "=B11*C11"
    ws['D11'].number_format = '#,##0.00'
    
    ws['A12'] = "Sound System Setup"
    ws['B12'] = 1
    ws['C12'] = 5000
    ws['C12'].number_format = '#,##0.00'
    ws['D12'] = "=B12*C12"
    ws['D12'].number_format = '#,##0.00'
    
    # Totals Section
    ws['C17'] = "Subtotal:"
    ws['C17'].font = Font(bold=True)
    ws['C17'].alignment = Alignment(horizontal="right")
    ws['D17'] = '=SUM(D11:D16)'
    ws['D17'].number_format = '#,##0.00'
    ws['D17'].font = Font(bold=True)
    
    ws['C18'] = "VAT (15%):"
    ws['C18'].alignment = Alignment(horizontal="right")
    ws['D18'] = '=D17*Settings!B4/100'
    ws['D18'].number_format = '#,##0.00'
    
    ws['C19'] = "TOTAL:"
    ws['C19'].font = Font(bold=True, size=13)
    ws['C19'].alignment = Alignment(horizontal="right")
    ws['D19'] = '=D17+D18'
    ws['D19'].number_format = '#,##0.00'
    ws['D19'].font = Font(bold=True, size=13)
    ws['D19'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['D19'].font = Font(bold=True, size=13, color="FFFFFF")
    
    # Payment Info
    ws['A21'] = "PAYMENT INFORMATION"
    ws['A21'].font = Font(bold=True, size=11)
    apply_header_style(ws, 21, 1, 2)
    
    ws['A22'] = "Bank: Sample Bank"
    ws['A23'] = "Account Name: Flash Lighting Services Ltd"
    ws['A24'] = "Account Number: 0123456789"
    
    # Footer
    ws['A26'] = "Thank you for your business!"
    ws['A26'].font = Font(italic=True, size=10)
    ws['A26'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A26:D26')
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # Set print area
    ws.print_area = 'A1:F26'


def create_audit_log_sheet(wb):
    """Create Audit Log sheet for tracking changes"""
    ws = wb.create_sheet("Audit Log")
    
    # Title
    ws['A1'] = "AUDIT LOG - TRANSACTION HISTORY"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "This sheet tracks all financial transactions for audit purposes"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = [
        "Timestamp", "Sheet", "Transaction Type", "Reference No.", 
        "Amount", "Description"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
    
    apply_header_style(ws, 4, 1, len(headers))
    
    # Sample audit entries
    ws['A5'] = datetime.now()
    ws['A5'].number_format = 'YYYY-MM-DD HH:MM:SS'
    ws['B5'] = "Invoices"
    ws['C5'] = "New Invoice Created"
    ws['D5'] = "INV-2025-001"
    ws['E5'] = 17250
    ws['E5'].number_format = '#,##0.00'
    ws['F5'] = "Sample Client Ltd - Event lighting setup"
    
    ws['A6'] = datetime.now()
    ws['A6'].number_format = 'YYYY-MM-DD HH:MM:SS'
    ws['B6'] = "Expenses"
    ws['C6'] = "New Expense Recorded"
    ws['D6'] = "TXN-001"
    ws['E6'] = 25000
    ws['E6'].number_format = '#,##0.00'
    ws['F6'] = "Equipment Purchase - Tech Supplies Ltd"
    
    # Instructions
    ws['A8'] = "Note: This log should be updated whenever financial transactions are created or modified."
    ws['A8'].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells('A8:F8')
    
    # Set column widths
    widths = [20, 15, 25, 18, 15, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'


def create_dashboard_sheet(wb):
    """Create Business Health Dashboard"""
    ws = wb.create_sheet("Dashboard")
    
    # Add company logo and branding
    logo_added = add_logo_to_sheet(ws, 'A1')
    
    # Title - adjust position if logo was added
    title_start = 'C1' if logo_added else 'A1'
    title_merge = 'C1:F1' if logo_added else 'A1:F1'
    
    ws[title_start] = COMPANY_NAME
    ws[title_start].font = Font(bold=True, size=18, color="FFFFFF")
    ws[title_start].fill = PatternFill(start_color=BRAND_COLOR_PRIMARY, end_color=BRAND_COLOR_PRIMARY, fill_type="solid")
    ws[title_start].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(title_merge)
    
    # Subtitle
    subtitle_cell = 'C2' if logo_added else 'A2'
    subtitle_merge = 'C2:F2' if logo_added else 'A2:F2'
    
    ws[subtitle_cell] = "BUSINESS HEALTH DASHBOARD"
    ws[subtitle_cell].font = Font(bold=True, size=12, color="FFFFFF")
    ws[subtitle_cell].fill = PatternFill(start_color=BRAND_COLOR_SECONDARY, end_color=BRAND_COLOR_SECONDARY, fill_type="solid")
    ws[subtitle_cell].alignment = Alignment(horizontal="center")
    ws.merge_cells(subtitle_merge)
    
    # Updated timestamp
    update_cell = 'C3' if logo_added else 'A3'
    update_merge = 'C3:F3' if logo_added else 'A3:F3'
    
    ws[update_cell] = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws[update_cell].alignment = Alignment(horizontal="center")
    ws[update_cell].font = Font(size=9, italic=True)
    ws.merge_cells(update_merge)
    
    # Key Metrics Section
    ws['A4'] = "KEY FINANCIAL METRICS"
    ws['A4'].font = Font(bold=True, size=13)
    ws.merge_cells('A4:F4')
    
    # Metric cards
    metrics = [
        ("Monthly Revenue", '=Revenue!L4', 'A6', 'B6'),
        ("Monthly Expenses", '=Expenses!J3', 'A7', 'B7'),
        ("Monthly Net Profit", '=B6-B7', 'A8', 'B8'),
        ("Profit Margin", '=IF(B6=0,0,B8/B6)', 'A9', 'B9'),
        ("", "", "", ""),
        ("YTD Revenue", '=Revenue!L5', 'A11', 'B11'),
        ("YTD Expenses", '=Expenses!J4', 'A12', 'B12'),
        ("YTD Net Profit", '=B11-B12', 'A13', 'B13'),
        ("YTD Profit Margin", '=IF(B11=0,0,B13/B11)', 'A14', 'B14'),
    ]
    
    for metric_name, formula, label_cell, value_cell in metrics:
        if metric_name:
            ws[label_cell] = metric_name
            ws[label_cell].font = Font(bold=True)
            ws[value_cell] = formula
            if "Margin" in metric_name:
                ws[value_cell].number_format = '0.00%'
            else:
                ws[value_cell].number_format = '#,##0.00'
            ws[value_cell].font = Font(bold=True, size=12)
            ws[value_cell].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Cash Position
    ws['D6'] = "Cash in Hand"
    ws['D6'].font = Font(bold=True)
    ws['E6'] = "='Cash Position'!B8"
    ws['E6'].number_format = '#,##0.00'
    ws['E6'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['D7'] = "Cash at Bank"
    ws['D7'].font = Font(bold=True)
    ws['E7'] = "='Cash Position'!B14"
    ws['E7'].number_format = '#,##0.00'
    ws['E7'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws['D8'] = "Total Cash"
    ws['D8'].font = Font(bold=True)
    ws['E8'] = "=E6+E7"
    ws['E8'].number_format = '#,##0.00'
    ws['E8'].font = Font(bold=True, size=12)
    ws['E8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Receivables and KPIs
    ws['D10'] = "Outstanding Invoices"
    ws['D10'].font = Font(bold=True)
    ws['E10'] = "='Trade Debtors'!J3"
    ws['E10'].number_format = '#,##0.00'
    ws['E10'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws['D11'] = "Total Stock Value"
    ws['D11'].font = Font(bold=True)
    ws['E11'] = "=Inventory!H12"
    ws['E11'].number_format = '#,##0.00'
    ws['E11'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws['D12'] = "VAT Payable"
    ws['D12'].font = Font(bold=True)
    ws['E12'] = "='Tax Summary'!B12"
    ws['E12'].number_format = '#,##0.00'
    ws['E12'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # Key Performance Indicators
    ws['A16'] = "KEY PERFORMANCE INDICATORS (KPIs)"
    ws['A16'].font = Font(bold=True, size=12)
    ws.merge_cells('A16:F16')
    
    # Days Sales Outstanding (DSO)
    ws['A17'] = "Days Sales Outstanding (DSO)"
    ws['A17'].font = Font(bold=True)
    ws['B17'] = '=IF(Revenue!L4=0,0,("Trade Debtors"!J3/(Revenue!L5/365)))'
    ws['B17'].number_format = '0.0 "days"'
    ws['B17'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Quick Ratio (Current Assets / Current Liabilities)
    ws['A18'] = "Quick Ratio"
    ws['A18'].font = Font(bold=True)
    ws['B18'] = '=IF("Balance Sheet"!C21=0,0,"Balance Sheet"!C9/"Balance Sheet"!C21)'
    ws['B18'].number_format = '0.00'
    ws['B18'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Conditional formatting for Quick Ratio (healthy > 1.0)
    ws.conditional_formatting.add('B18:B18',
        CellIsRule(operator='greaterThan', formula=['1'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    
    ws.conditional_formatting.add('B18:B18',
        CellIsRule(operator='lessThan', formula=['1'], 
                   stopIfTrue=True, 
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    # Gross Profit Margin
    ws['D17'] = "Gross Profit Margin"
    ws['D17'].font = Font(bold=True)
    ws['E17'] = '="Profit & Loss"!D30'
    ws['E17'].number_format = '0.00%'
    ws['E17'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Inventory Turnover
    ws['D18'] = "Total Assets"
    ws['D18'].font = Font(bold=True)
    ws['E18'] = '="Balance Sheet"!C15'
    ws['E18'].number_format = '#,##0.00'
    ws['E18'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Revenue vs Expense Trend
    ws['A20'] = "REVENUE VS EXPENSE TREND (Monthly)"
    ws['A20'].font = Font(bold=True, size=12)
    ws.merge_cells('A20:F20')
    
    ws['A21'] = "Month"
    ws['B21'] = "Revenue"
    ws['C21'] = "Expenses"
    ws['D21'] = "Net Profit"
    apply_header_style(ws, 21, 1, 4)
    
    months = ["January", "February", "March", "April", "May", "June", 
              "July", "August", "September", "October", "November", "December"]
    
    for idx, month in enumerate(months, start=22):
        month_num = idx - 21
        ws[f'A{idx}'] = month
        ws[f'B{idx}'] = f'=SUMIFS(Revenue!H:H,Revenue!A:A,">="&DATE(YEAR(TODAY()),{month_num},1),Revenue!A:A,"<"&DATE(YEAR(TODAY()),{month_num}+1,1))'
        ws[f'B{idx}'].number_format = '#,##0.00'
        ws[f'C{idx}'] = f'=SUMIFS(Expenses!E:E,Expenses!A:A,">="&DATE(YEAR(TODAY()),{month_num},1),Expenses!A:A,"<"&DATE(YEAR(TODAY()),{month_num}+1,1))'
        ws[f'C{idx}'].number_format = '#,##0.00'
        ws[f'D{idx}'] = f'=B{idx}-C{idx}'
        ws[f'D{idx}'].number_format = '#,##0.00'
        
        # Conditional formatting for profit/loss
        ws.conditional_formatting.add(f'D{idx}:D{idx}',
            CellIsRule(operator='greaterThan', formula=['0'], 
                       stopIfTrue=True, 
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
        
        ws.conditional_formatting.add(f'D{idx}:D{idx}',
            CellIsRule(operator='lessThan', formula=['0'], 
                       stopIfTrue=True, 
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    
    # Create chart
    chart = LineChart()
    chart.title = "Revenue vs Expense Trend"
    chart.style = 12
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Month"
    chart.height = 10
    chart.width = 20
    
    # Add data
    data = Reference(ws, min_col=2, min_row=21, max_row=33, max_col=3)
    cats = Reference(ws, min_col=1, min_row=22, max_row=33)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "F22")
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 3


def main():
    """Main function to generate the workbook"""
    print("\n" + "="*70)
    print("  PROFESSIONAL BOOKKEEPING SYSTEM GENERATOR")
    print("  QuickBooks Edition with Company Branding")
    print("="*70 + "\n")
    
    print(f"📋 Company: {COMPANY_NAME}")
    print(f"📧 Contact: {COMPANY_EMAIL} | {COMPANY_PHONE}")
    
    # Check for logo
    if os.path.exists(LOGO_FILENAME):
        print(f"✅ Logo found: {LOGO_FILENAME}")
        print(f"   Size: {LOGO_WIDTH}x{LOGO_HEIGHT} pixels")
    else:
        print(f"⚠️  Logo not found: {LOGO_FILENAME}")
        print(f"   Place your logo file in this folder to include it")
        print(f"   Supported formats: PNG, JPG, GIF")
    
    print("\n🔨 Generating workbook...")
    
    wb = create_workbook()
    
    filename = "Event_Lighting_Bookkeeping.xlsx"
    wb.save(filename)
    
    print(f"✓ Workbook created successfully: {filename}")
    print("\n" + "="*60)
    print("PROFESSIONAL FEATURES INCLUDED:")
    print("="*60)
    print("\n📊 Financial Reports:")
    print("  ✓ Dashboard (Business Health Overview)")
    print("  ✓ Profit & Loss Statement")
    print("  ✓ Balance Sheet")
    print("  ✓ Cash Position Statement")
    print("  ✓ Tax Summary Report")
    print("  ✓ Bank Reconciliation")
    
    print("\n👥 Management:")
    print("  ✓ Customer Database (with credit limits & tracking)")
    print("  ✓ Vendor/Supplier Database")
    print("  ✓ Inventory Management (with stock valuation)")
    
    print("\n💼 Operations:")
    print("  ✓ Invoice Management (with auto-calculations)")
    print("  ✓ Revenue Tracking (auto-sync from invoices)")
    print("  ✓ Expense Tracking (with categories)")
    print("  ✓ Trade Debtors (Accounts Receivable)")
    print("  ✓ Professional Invoice Template (print-ready)")
    
    print("\n🔒 Professional Features:")
    print("  ✓ Data Validation (dropdown lists for consistency)")
    print("  ✓ Conditional Formatting (visual alerts)")
    print("  ✓ Audit Log (transaction history)")
    print("  ✓ Cross-sheet formulas (auto-updating)")
    print("  ✓ Settings Sheet (centralized configuration)")
    
    print("\n" + "="*60)
    print("SHEETS CREATED:")
    print("="*60)
    visible_sheets = []
    hidden_sheets = []
    for sheet in wb.sheetnames:
        if wb[sheet].sheet_state == 'hidden':
            hidden_sheets.append(sheet)
        else:
            visible_sheets.append(sheet)
    
    for sheet in visible_sheets:
        print(f"  📄 {sheet}")
    
    print("\n🔒 Hidden Sheets (for advanced users):")
    for sheet in hidden_sheets:
        print(f"  🔐 {sheet}")
    
    print("\n" + "="*60)
    print("\n✨ Your professional bookkeeping system is ready!")
    print("\n📝 Next Steps:")
    print("  1. Open the file and start with the Dashboard")
    print("  2. Customize Settings sheet (unhide it first)")
    print("  3. Add your customers and vendors")
    print("  4. Start creating invoices and tracking expenses")
    print("  5. Financial reports update automatically!\n")


if __name__ == "__main__":
    main()
